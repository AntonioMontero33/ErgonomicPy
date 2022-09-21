import openpyxl
from openpyxl import load_workbook
import smtplib

class Paciente:
  def __init__(self, dni, nombre, empresa, company, tipo, fecha, correoDelDoctor, archivo):
    self.dni = dni
    self.nombre = nombre
    self.empresa = empresa
    self.company = company
    self.tipo = tipo
    self.fecha = fecha
    self.correoDelDoctor = correoDelDoctor
    self.archivo = archivo


wb = load_workbook(filename = 'try.xlsx')
ws = wb["Datos"]  
    
for i in range(2,ws.max_row):
  nombres1 = ws[f'A{i}']
  archivos = ws[f'B{i}']
  if nombres1.value != None:
    lsNombre = nombres1.value.split("-")
    for i in range(0,len(lsNombre)):
      lsNombre[i]=lsNombre[i].strip()
    archivoLast = nombres1.value + ".pdf" 
    p2 = Paciente(lsNombre[0], lsNombre[1], lsNombre[2], lsNombre[3], lsNombre[4], lsNombre[5], archivos.value, archivoLast)
    
    

print(p2.dni)
print(p2.nombre)
print(p2.empresa)
print(p2.company)
print(p2.tipo)
print(p2.fecha)
print(p2.correoDelDoctor)
print(p2.archivo)
    


